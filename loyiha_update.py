from openpyxl import Workbook, load_workbook
from datetime import datetime


class Shop:
    def __init__(self, file_path):
        self.file_path = file_path
        self.input_field = dict()
        self.data = dict()
        self.read_xlsx()

    def read_multe_sheet(self):
        pass

    def read_xlsx(self):
        wb = load_workbook(self.file_path)
        ws = wb.active
        for i in range(1, ws.max_column + 1):
            self.input_field[ws.cell(row=1, column=i).value] = ws.cell(row=2, column=i).value

        for i in range(1, ws.max_column + 1):
            self.data[ws.cell(row=1, column=i).value] = list()
            for j in range(3, ws.max_row + 1):
                self.data.get(ws.cell(row=1, column=i).value).append(ws.cell(row=j, column=i).value)

    def add_data(self, name, amount, price, date):
        wb = load_workbook(self.file_path)
        ws = wb.active
        ws["A1"] = "Name"
        ws["B1"] = "Amount"
        ws["C1"] = "Price"
        ws["D1"] = "Date"
        max_row = ws.max_row
        ws.cell(row=max_row + 1, column=1, value=name)
        ws.cell(row=max_row + 1, column=2, value=amount)
        ws.cell(row=max_row + 1, column=3, value=price)
        ws.cell(row=max_row + 1, column=4, value=date)
        wb.save(self.file_path)
        self.read_xlsx()

    def sell_data(self, product_name, selling_amount):
        wb = load_workbook(self.file_path)
        ws = wb.active
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == product_name:
                current_row = row
                current_amount = ws.cell(row=current_row, column=2).value
                if current_amount >= selling_amount:
                    ws.cell(row=current_row, column=2, value=current_amount - selling_amount)
                    ws.cell(row=current_row, column=5, value=selling_amount)
                    ws.cell(row=current_row, column=6, value=datetime.now())
                    wb.save(self.file_path)
                    self.read_xlsx()  
                    print("Muvaffaqiyatli sotildi")
                    return
                else:
                    print("Mahsulot sotilmadi")
                    return
            print("Mahsulot topilmadi")


    def report(self, report_file_path):
        if "Name" not in self.data:
            print("Dasturda xatolik")
            return

        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Amount", "Price", "Date Sold"])

        for i in range(len(self.data["Name"])):
            ws.append([
                self.data["Name"][i],
                self.data["Amount"][i],
                self.data["Price"][i],
                self.data["Date"][i]
            ])

        wb.save(report_file_path)
        print("Report generated successfully.")

    def main(self):
        while True:
            print("1. Add Product")
            print("2. Sell Product")
            print("3. Generate Report")
            print("4. Exit")
            choice = int(input("Enter your choice: "))
            if choice == 1:
                name = input("Enter product name: ")
                amount = int(input("Enter amount: "))
                price = float(input("Enter price: "))
                date = input("Enter date (YYYY-MM-DD): ")
                self.add_data(name, amount, price, date)
            elif choice == 2:
                product_name = input("Enter product name to sell: ")
                selling_amount = float(input("Enter selling amount: "))
                self.sell_data(product_name, selling_amount)
            elif choice == 3:
                file_path = "data.xlsx"
                self.report(file_path)
            elif choice == 4:
                break
            else:
                print("Invalid choice.")


file_path = "data.xlsx"
shop = Shop(file_path)
shop.main()
